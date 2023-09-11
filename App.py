# from customtkinter import *
from tkcalendar import DateEntry
from tkinter import Menu, MULTIPLE, SINGLE,Tk,Label,Frame,IntVar,StringVar,DoubleVar,Listbox,END,Toplevel,Scrollbar,font, _setit, messagebox
from models import *
from tkinter import ttk
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from sqlalchemy import asc
import random
from ttkwidgets.autocomplete import AutocompleteCombobox

# import models
# import win32print

# height = 450
# width = 600
root = Tk()
root.config(background='#ffffff')
root.geometry("1000x800")
root.state('zoomed')
root.title("Inventory Management System")
w = root.winfo_screenmmwidth()
h = root.winfo_screenmmheight()
style = ttk.Style()

navbar = Menu(root, font=('Arial', 20))
root.config(menu=navbar)



# Menu Tab Functions =============================================
def purchase():
    hideAllFrame()
    frame1.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def sales():
    hideAllFrame()
    frame2.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def cold_facility():
    hideAllFrame()
    frame3.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def party_name():
    hideAllFrame()
    frame4.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def report_screen():
    hideAllFrame()
    report_frame.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def WeightPending():
    hideAllFrame()
    weight_frame.grid(row=0,column=0,padx=10,pady=10,sticky="nsew")

def hideAllFrame():
    frame1.grid_remove()
    frame2.grid_remove()
    frame3.grid_remove()
    frame4.grid_remove()
    report_frame.grid_remove()
    weight_frame.grid_remove()
    rootLable.grid_remove()

navbar.add_command(label="Report",command=report_screen)
navbar.add_command(label="Purchase",command=purchase)
navbar.add_command(label="Sales", command=sales)
navbar.add_command(label="Cold Facility", command=cold_facility)
navbar.add_command(label="Party Name", command=party_name)
navbar.add_command(label="Pending Weight Lots", command=WeightPending)

# End of Menu Tab Functions ==========================================
# Trail screen================================================================

# trail =14-int(str(date.today()).split("-")[2])

warn = ttk.Label(master=root,text="",font=("times new roman",14))
warn.grid(row=1,column=0,columnspan=5,pady=10)

# if(trail):
#   warn.config(text=f"{trail} Days Trail period Remaining. Please Purchase Full Version")
# else:
#   warn.config(text="Trail period is OVER. Please Purchase Full Version")
#   navbar.entryconfigure("Purchase",state="disabled")
#   navbar.entryconfigure("Sales",state="disabled")
#   navbar.entryconfigure("Cold Facility",state="disabled")
#   navbar.entryconfigure("Party Name",state="disabled")


#=============================================================================

# Functions=================================================================
def getColdFacilityOptions():
    global coldFacilityOptions
    coldFacilityOptions = ["0 None"]
    data = session.query(ColdFacility).all()
    for item in data:
        coldFacilityOptions.append(item)

def getpartyNameOptions():
    global partyNameOptions
    data = session.query(PartyName).order_by(asc(PartyName.Name)).all()
    for item in data:
        partyNameOptions.append(str(item))

def clear_partyName_option():
    global defaultPartyOption
    firmNameVar.set(defaultPartyOption)
    firmNameVarFilter.set(defaultPartyOption)
    sale_soldToNameVarFilter.set(defaultPartyOption)

def clear_coldFacility_option():
    global defaultColdFacilityOption
    coldVar.set(defaultColdFacilityOption)
    firmNameVarFilter.set(defaultColdFacilityOption)
    
def purchaseListSelect(event):
    try:
        global idPurchase
        index = event.widget.curselection()
        text = str(purchaseList.get(index))
        datalist = text.split(sep=" | ")
        idPurchase=int(datalist[0].split(':')[1])
        idParty = int(datalist[2].split(':')[1])
        idCold=datalist[8].split(':')[1]
        data = session.query(Purchase).filter(Purchase.id==idPurchase).first()
        data1 = session.query(PartyName).filter(PartyName.id==idParty).first()
        coldData = session.query(ColdFacility).filter(ColdFacility.id==idCold).first()
        if(data):
            clear_purchase_input()
            auctionDateEntry.set_date(data.AuctionDate)
            firmNameVar.set(data1)
            markingIdEntry.insert(0,data.MarkingID)
            boxEntry.insert(0,data.Box)
            auctionRateEntry.insert(0,data.AuctionRate)
            weightEntry.insert(0,data.weight)
            coldVar.set(coldData)

            addSalesbtn.config(state="normal")

            purchase_info.delete(0,END)
            if(data.weight==float(0)):               
                purchase_info.insert(END,f"LOT Weight Pending")
            else:
                purchase_info.insert(END,f"Total Weight of LOT {data.weight} Kg")
            if(data.Box>0):
                purchase_info.insert(END,f"{data.Box} Box remains to sold")
            purchase_info.insert(END,f"Purchased from {data1.Name}")
    

    except:
        pass

def salesListSelect(event):
    try:
        global idSales
        index = event.widget.curselection()
        text = str(salesList.get(index))
        datalist = text.split(sep=" | ")
        idSales=int(datalist[0].split(': ')[1])

        data = session.query(Sell).filter(Sell.id==idSales).first()
        data1 = session.query(Purchase).filter(Purchase.id == data.purchase_id).first()
        data2 = session.query(PartyName).filter(PartyName.id == str(data.SellTo).split(' | ')[0]).first()
        coldData = session.query(ColdFacility).filter(ColdFacility.id==data1.coldfacility_id).first()
        if(data):
            clear_sales_input()
            sale_soldToNameVarFilter.set(data2)
            sale_markingIdEntryFilter.insert(0,str(data.MarkingID))
            # sale_coldVarFilter.set(coldFacilityOptions[data1.coldfacility_id])
            d =""
            dd=""
            c=coldData.Name if coldData else "Cold Facility not found"
            if(data.Dispatched):
                d= ""
                dd=data.DispatchDate
            else: 
                d=" NOT"
                dd="NONE"

            sale_info.delete(0,END)
            sale_info.insert(END,f"Kept in {c} Facility")
            sale_info.insert(END,f"LOT is{d} Dispatched")
            sale_info.insert(END,f"LOT Dispatched Date: {dd}")
        
    except:
        pass

def getPendingWeight():
    try:
        weight_pen = session.query(Purchase).filter(Purchase.weight < 1)        
        if(weight_pen.count()>0):
            # print(len(data.all()))
            # for i in data.all():
            #     print('data    ',i)
            # all_coldFacility_data.delete(0,END)
            weight_tree.delete(*weight_tree.get_children())
            # all_coldFacility_data.insert(END,f"Total LOTS which Not Dispatched = {data.count()}")
            ci=1
            for i in weight_pen.all():
                l = str(i).split(' | ')
                partyId = l[2].split(':')[1] if (l[2] and len(l[2].split(':'))>1) else None
                firmName = "No Name"
                if partyId:
                    pname = session.query(PartyName).filter(PartyName.id == partyId).first()
                    if pname and pname.Name:
                        firmName=pname.Name
                # t_insert = "{:^15} | {:^25} | {:^15} | {:^10} | {:^15} | {:^10} | {:^25} | {:^10}"
                # t_insert = t_insert.format(
                #     l[1].split(':')[1],
                #     pname.Name,l[3].split(':')[1],
                #     l[12].split(': ')[1],
                #     l[5].split(':')[1],
                #     l[7].split(':')[1],
                #     l[11],
                #     idCold)
                # all_coldFacility_data.insert(END,t_insert)
                weight_tree.insert("", 'end', text =f"{ci}",
                                             values =(
                    l[0].split(':')[1] if (l[0] and len(l[0].split(':'))>1) else "None",
                    l[1].split(':')[1] if (l[1] and len(l[1].split(':'))>1) else "None",
                    firmName,
                    l[3].split(':')[1] if (l[3] and len(l[3].split(':'))>1) else "None",
                    l[4].split(':')[1] if (l[4] and len(l[4].split(':'))>1) else "None",
                    l[5].split(':')[1] if (l[5] and len(l[5].split(':'))>1) else "None",
                    l[6].split(':')[1] if (l[6] and len(l[6].split(':'))>1) else "None",
                    l[7].split(':')[1] if (l[7] and len(l[7].split(':'))>1) else "None",
                    l[8].split(':')[1] if (l[8] and len(l[8].split(':'))>1) else "None"
                    ))
                ci+=1                
                # all_coldFacility_data.insert(END,item)
        
        # coldFacility_data_scrlbar = ttk.Scrollbar(frame3,orient ="vertical",command = all_coldFacility_data.yview)
        # coldFacility_data_scrlbar.grid(row=1,column=4,rowspan=14,sticky="nsw")
        # all_coldFacility_data.configure(yscrollcommand = coldFacility_data_scrlbar.set)


    except:
        pass


def coldListSelect(event):
    try:
        global idCold
        index = event.widget.curselection()
        text = str(coldFacilityList.get(index))
        datalist = text.split(sep=" | ")
        id = datalist[0]
        n = datalist[1]       
        idCold = int(id)
        coldFacilityEntry.delete(0,END)
        coldFacilityEntry.insert(0,n)
        data = session.query(Purchase,Sell).filter(Purchase.id==Sell.purchase_id,
                                                   Sell.Dispatched==0,
                                                   Purchase.coldfacility_id==id)
        ci = 1
        if(data):
            # print(len(data.all()))
            # for i in data.all():
            #     print('data    ',i)
            # all_coldFacility_data.delete(0,END)
            all_coldFacility_data.delete(*all_coldFacility_data.get_children())
            # all_coldFacility_data.insert(END,f"Total LOTS which Not Dispatched = {data.count()}")

            for i in data.all():
                # print(i)
                l = str(i).split(' | ')
                pname = session.query(PartyName).filter(PartyName.id == int(l[2].split(':')[1])).first()
                firmName = "No Name"
                if pname and pname.Name:
                    firmName=pname.Name
                # t_insert = "{:^15} | {:^25} | {:^15} | {:^10} | {:^15} | {:^10} | {:^25} | {:^10}"
                # t_insert = t_insert.format(
                #     l[1].split(':')[1],
                #     pname.Name,l[3].split(':')[1],
                #     l[12].split(': ')[1],
                #     l[5].split(':')[1],
                #     l[7].split(':')[1],
                #     l[11],
                #     idCold)
                # all_coldFacility_data.insert(END,t_insert)

                all_coldFacility_data.insert("", 'end', text =f"{ci}",
                                             values =(
                    l[1].split(':')[1],
                    firmName,
                    l[3].split(':')[1],
                    l[12].split(': ')[1],
                    l[5].split(':')[1],
                    l[7].split(':')[1],
                    l[11],
                    idCold
                    ))
                ci+=1
                # all_coldFacility_data.insert(END,item)
                
        data1 = session.query(Purchase).filter(Purchase.coldfacility_id==id,Purchase.Box>0)
        # all_coldFacility_data.insert(END,"=========================================================")
        # all_coldFacility_data.insert(END,f"Total LOTS which Not Sold = {data1.count()}")
        if(data1):
            for item in data1:
                l = str(item).split(' | ')
                pname = session.query(PartyName).filter(PartyName.id == int(l[2].split(':')[1])).first()
                firmName = "No Name"
                if pname and pname.Name:
                    firmName=pname.Name

                all_coldFacility_data.insert("", 'end', text =f"{ci}",
                                            values =(
                    l[1].split(':')[1],
                    firmName,
                    l[3].split(':')[1],
                    l[4].split(':')[1],
                    l[5].split(':')[1],
                    l[7].split(':')[1],
                    "None",
                    l[8].split(':')[1]))
                ci+=1

                # all_coldFacility_data.insert(END,f"{l[1].split(':')[1]} | Purchase From:{pname.Name} | {l[3]} | Sold To: None | {l[4]} | {l[5]}")                   

        coldFacility_data_scrlbar = ttk.Scrollbar(frame3,orient ="vertical",command = all_coldFacility_data.yview)
        coldFacility_data_scrlbar.grid(row=1,column=4,rowspan=14,sticky="nsw")
        all_coldFacility_data.configure(yscrollcommand = coldFacility_data_scrlbar.set)


    except:
        pass

def partyListSelect(event):
    try:
        global idParty
        index = event.widget.curselection()
        text = str(partyNameList.get(index))
        datalist = text.split(sep=" | ")
        id = datalist[0]
        n = datalist[1]       
        idParty = int(id)
        partyNameEntry.delete(0,END)
        partyNameEntry.insert(0,n)
        data1 = session.query(Sell).filter(Sell.SellTo==text)
        data = session.query(Purchase).filter(Purchase.FirmName==idParty)
        i=2
        if(data):
            all_partyName_data.delete(0,END)
            all_partyName_data.insert(1,f"Total Purchase {data.count()}")           
            for item in data:
                l = str(item).split(' | ')
                all_partyName_data.insert(i,item)
                i+=1
        if(data1):
            i+=1
            all_partyName_data.insert(i,f"Total Sales {data1.count()}")
            i+=1 
            for item in data1:
                l = str(item).split(' | ')
                all_partyName_data.insert(i,item)
                i+=1
        
    except:
        pass

def clearPurchase():
    auctionDateEntry.set_date(date.today())
    clear_partyName_option()
    markingIdEntry.delete(0,END)
    boxEntry.delete(0,END)
    auctionRateEntry.delete(0,END)
    weightEntry.delete(0,END)
    weightEntry.insert(0,"0.0")
    clear_coldFacility_option()

def clear_purchase_input():
    clear_partyName_option()
    markingIdEntry.delete(0,END)
    boxEntry.delete(0,END)
    auctionRateEntry.delete(0,END)
    weightEntry.delete(0,END)
    clear_coldFacility_option()

def calculate_rate(auction_rate_var, rate_var):
    try:
        auction_rate = float(auction_rate_var.get())
        rate = auction_rate + (0.025 * auction_rate)
        rate_var.set(round(rate, 1))
    except ValueError:
        rate_var.set("Invalid Input")

def add_multiple_purchases():
    # Create a new dialog window
    dialog = Toplevel(root)
    dialog.title("Add Multiple Purchases")

    # Create labels and entry fields for common details (auction date and firm)
    ttk.Label(dialog, text="Auction Date:").grid(row=0, column=0)
    auction_date_entry = DateEntry(dialog, selectmode = 'day', date_pattern="dd-mm-yyyy")
    auction_date_entry.grid(row=0, column=1)
    getpartyNameOptions()
    ttk.Label(dialog, text="Firm Name:").grid(row=0, column=3)
    firm_name_var = StringVar()
    print(partyNameOptions)
    firm_name_dropdown = AutocompleteCombobox(dialog, textvariable=firm_name_var, completevalues=partyNameOptions)
    # firm_name_dropdown['values'] = ('Firm A', 'Firm B', 'Firm C')  # Replace with your firm names
    firm_name_dropdown.grid(row=0, column=4)

    # Create labels for purchase details
    ttk.Label(dialog, text="Marking ID").grid(row=2, column=0)
    ttk.Label(dialog, text="Box").grid(row=2, column=1)
    ttk.Label(dialog, text="Auction Rate").grid(row=2, column=2)
    ttk.Label(dialog, text="Rate").grid(row=2, column=3)
    ttk.Label(dialog, text="Net Weight").grid(row=2, column=4)
    ttk.Label(dialog, text="Cold Facility").grid(row=2, column=5)

    # Button to delete the last added purchase row
    def delete_purchase_row():
        if purchase_rows:
            last_row = purchase_rows.pop()
            for widget in last_row:
                try:
                    widget.grid_remove()
                except:
                    pass

            if not purchase_rows:
                # Disable the delete button if there are no rows left
                delete_button.config(state='disabled')

    delete_button = ttk.Button(dialog, text="Delete Last Purchase", command=delete_purchase_row, state='disabled')
    delete_button.grid(row=4, column=6)

    # Function to add a new row for purchase details
    def add_purchase_row():
        common_auction_date = auction_date_entry.get()
        common_firm_name = firm_name_var.get()

        if not common_auction_date or not common_firm_name:
            messagebox.showerror("Error", "First fill Auction Date and Firm Name.", parent=dialog)
            return

        marking_id_entry = ttk.Entry(dialog)
        marking_id_entry.grid(row=3 + len(purchase_rows), column=0)
        
        box_entry = ttk.Entry(dialog)
        box_entry.grid(row=3 + len(purchase_rows), column=1)

        auction_rate_var = StringVar()
        auction_rate_entry = ttk.Entry(dialog, textvariable=auction_rate_var)
        auction_rate_entry.grid(row=3 + len(purchase_rows), column=2)
        
        net_weight_entry = ttk.Entry(dialog)
        net_weight_entry.grid(row=3 + len(purchase_rows), column=4)

        cold_facility_var = StringVar()
        getColdFacilityOptions()
        cold_facility_dropdown = ttk.OptionMenu(dialog,cold_facility_var,defaultColdFacilityOption,*coldFacilityOptions)
        # cold_facility_dropdown = ttk.Combobox(dialog, textvariable=cold_facility_var)
        # cold_facility_dropdown['values'] = ('Cold A', 'Cold B', 'Cold C')  # Replace with your cold facilities
        cold_facility_dropdown.grid(row=3 + len(purchase_rows), column=5)

        rate_var = StringVar()
        rate_entry = ttk.Entry(dialog, textvariable=rate_var, state='readonly')
        rate_entry.grid(row=3 + len(purchase_rows), column=3)

        purchase_rows.append((marking_id_entry, box_entry, auction_rate_entry, rate_entry, rate_var, net_weight_entry, cold_facility_dropdown, cold_facility_var))

        # Enable the delete button
        delete_button.config(state='normal')

    purchase_rows = []  # List to store dynamically added purchase rows

    # Add the first row for purchase details
    # add_purchase_row()

    # Button to add more purchase rows
    ttk.Button(dialog, text="Add Purchase", command=add_purchase_row).grid(row=3, column=6)



    # Function to save the purchases to the database
    def save_purchases():
        common_auction_date = auction_date_entry.get()
        common_firm_name = firm_name_var.get()

        if not common_auction_date or not common_firm_name:
            messagebox.showerror("Error", "Auction Date and Firm Name are required.", parent=dialog)
            return

        for i, (marking_id_entry, box_entry, auction_rate_entry, rate_entry, rate_var, net_weight_entry, cold_facility_dropdown, cold_facility_var) in enumerate(purchase_rows):
            marking_id = marking_id_entry.get()
            box = box_entry.get()
            auction_rate = auction_rate_entry.get()
            net_weight = net_weight_entry.get()
            cold_facility = cold_facility_var.get()
            rate = rate_entry.get()

            if not marking_id or not box or not auction_rate or not net_weight or not cold_facility:
                messagebox.showerror("Error", "All fields are required.", parent=dialog)
                return

            # Calculate the rate if auction_rate is valid
            try:
                auction_rate = float(auction_rate)
                rate = auction_rate + (0.025 * auction_rate)
                rate_var.set(round(rate, 1))
            except ValueError:
                rate_var.set("Invalid Input")

            # Insert the purchase details into the database using common_auction_date, common_firm_name,
            # marking_id, box, auction_rate, net_weight, cold_facility, and rate

            # Example: You can use a database library like SQLite or any other database you prefer
            # to insert the data into the database.

            # After inserting, you can provide feedback to the user about the success of the operation.

            # For simplicity, this example just prints the details to the console.
            print(f"Purchase {i + 1}:")
            print(f"Auction Date: {common_auction_date}")
            print(f"Firm Name: {common_firm_name}")
            print(f"Marking ID: {marking_id}")
            print(f"Box: {box}")
            print(f"Auction Rate: {auction_rate}")
            print(f"Net Weight: {net_weight}")
            print(f"Cold Facility: {cold_facility}")
            print(f"Rate: {rate}")
            print("\n")
            try:
                p = Purchase(AuctionDate=auction_date_entry.get_date(), 
                        FirmName=int(common_firm_name.split(" ")[0]), 
                        MarkingID=marking_id,
                        Box=box,
                        AuctionRate=float(auction_rate),
                        Rate=float(rate),
                        weight=float(net_weight),
                        coldfacility_id = int(cold_facility.split(" ")[0]))
                        
                session.add(p)
            except:
                messagebox.showerror("Error", f"Error occured adding {i+1} row", parent=dialog)
                pass

        session.commit()
        messagebox.showinfo("Success", "Added data", parent=dialog)
        # Close the dialog after saving the purchases
        getPurchase()
        dialog.destroy()

    # Button to save the purchases to the database
    ttk.Button(dialog, text="Save Purchases", command=save_purchases).grid(row=5, column=6)

def addPurchase():
    if(auctionDateEntry != "" and firmNameVar.get() != "" and markingidPurchase.get() != "" and auctionRateVar != 0):
        p = Purchase(AuctionDate=auctionDateEntry.get_date(), 
                     FirmName=int(firmNameVar.get().split(" ")[0]), 
                     MarkingID=markingidPurchase.get(),
                     Box=boxVar.get(),
                     AuctionRate=float(auctionRateVar.get()),
                     Rate=float(rateVar.get()),
                     weight=float(weightVar.get()),
                     coldfacility_id = int(coldVar.get().split(" ")[0]))
                     
        session.add(p)
        session.commit()
        clearPurchase()
        clear_purchase_input()
        getPurchase()
        status1["text"] = "Status: Product Added successfully"
        status1.config(background='green',foreground='white')
        status1.after(10000,lambda:status1.config(text="Status:",background='white',foreground='black'))
    else:
        status1["text"] = "Status: Product Not Added __Press Clear Button__"
        status1.config(background='red',foreground='white')
        status1.after(10000,lambda:status1.config(text="Status:",background='white',foreground='black'))

def editPurchase():
    if(firmNameVar.get() != "" and markingidPurchase.get() != ""):
        try:
            data = session.query(Purchase).filter(Purchase.id == idPurchase).first()
            if(data):
                data.AuctionDate = auctionDateEntry.get_date()
                data.FirmName=int(firmNameVar.get().split(" | ")[0])
                data.MarkingID=markingidPurchase.get()
                data.Box=boxVar.get()
                data.AuctionRate=float(auctionRateVar.get())
                data.Rate=float(rateVar.get())
                data.weight=weightVar.get()
                if len(coldVar.get().split(" | "))>1:
                    data.coldfacility_id=int(coldVar.get().split(" | ")[0])
                session.commit()
                clearPurchase()
                getPurchase()
                status1["text"] = "Status: Product Edit successfully"
                status1.config(background='green',foreground='white')
                status1.after(3000,lambda:status1.config(text="Status:",background='white',foreground='black'))
        except:
            status1["text"] = "Status: Fill all data, and try again!"
            status1.config(background='red',foreground='white')
            status1.after(3000,lambda:status1.config(text="Status:",background='white',foreground='black'))
        
def deletePurchase():
    if(firmNameVar.get() != "" and markingidPurchase.get() != ""):
        data = session.query(Purchase).filter(Purchase.id == idPurchase).first()
        if(data):
            session.delete(data)
            session.commit()
            clearPurchase()
            getPurchase()
            status1["text"] = "Status: Product deleted successfully"
            status1.config(background='green',foreground='white')
            status1.after(10000,lambda:status1.config(text="Status:",background='white',foreground='black'))
    
def clear_sales_input():
    clear_partyName_option()
    sale_markingIdEntryFilter.delete(0,END)

def addSales():
    popup=Toplevel(frame1,padx=30,pady=30)
    popup.title("Add Sales To")
    # data = session.query(Purchase).filter(Purchase.id==idPurchase).first()

    salesToNameVar = StringVar()
    salesToNameLabel = ttk.Label(popup,text="Sold To",anchor="w")
    salesToNameLabel.grid(row=0,column=0,padx=10,pady=10,sticky="w")
    # salesToNameEntry = ttk.OptionMenu(popup,salesToNameVar,defaultPartyOption,*partyNameOptions)
    salesToNameEntry =  AutocompleteCombobox(popup, textvariable=salesToNameVar, completevalues=partyNameOptions)
    salesToNameVar.set(defaultPartyOption)
    salesToNameEntry.grid(row=0,column=1,padx=10,pady=10)


    salesToBoxVar = IntVar()
    salesToBoxLabel = ttk.Label(popup,text="Box",anchor="w")
    salesToBoxLabel.grid(row=1,column=0,padx=10,pady=10,sticky="w")
    salesToBoxEntry = ttk.Entry(popup,textvariable=salesToBoxVar)
    salesToBoxEntry.grid(row=1,column=1,padx=10,pady=10)
    salesToBoxEntry.delete(0,END)
    salesToBoxEntry.insert(0,str(boxVar.get()))

    def addSalesTo():
        data = session.query(Purchase).filter(Purchase.id==idPurchase).first()
        # print(f"{salesToBoxVar.get()}=========================")
        if(salesToNameEntry != ""):
            s = Sell(MarkingID = data.MarkingID,
                     SellTo = salesToNameVar.get(),
                     Box = salesToBoxVar.get(),
                     purchase_id = data.id)
            session.add(s)
            data.Box = int(data.Box) - int(salesToBoxVar.get())
            session.commit()
            getSales()
            getPurchase()
            salesToNameVar.set(defaultPartyOption)
            salesToBoxEntry.delete(0,END)
            statusSalesTo["text"] = "Status: Sales Added Successfully"
            statusSalesTo.config(background='green',foreground='white')
            statusSalesTo.after(10000,lambda:statusSalesTo.config(text="Status:",background='white',foreground='black'))


    salesToAddbtn = ttk.Button(popup,text="ADD sales",command=addSalesTo)
    salesToAddbtn.grid(row=2,column=0,padx=20,pady=20,sticky="w")

    statusSalesTo = ttk.Label(popup,text="Status: ",anchor="w")
    statusSalesTo.grid(row=3,column=0,sticky="we",pady=5,columnspan=3)

def deleteSales():
    print('ssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssssss')
    data = session.query(Sell).filter(Sell.id == idSales).first()
    if(data):
        session.delete(data)
        session.commit()
        getSales()
        sale_status["text"] = "Status: Sales deleted successfully"
        sale_status.config(background='green',foreground='white')
        sale_status.after(10000,lambda:sale_status.config(text="Status:",background='white',foreground='black'))

def editSales():
    if(sale_soldToNameVarFilter.get() != defaultPartyOption and sale_markingidPurchaseFilter.get() != ""):       
        data = session.query(Sell).filter(Sell.id == idSales).first()
        if(data):
            data.SellTo = sale_soldToNameVarFilter.get().split(" ")[0]
            data.MarkingID=sale_markingidPurchaseFilter.get()
            session.commit()
            clearPurchase()
            getSales()
            sale_status["text"] = "Status: Sales Edit successfully"
            sale_status.config(background='green',foreground='white')
            sale_status.after(10000,lambda:sale_status.config(text="Status:",background='white',foreground='black'))

def dispatchedSales():
    popup=Toplevel(master=frame2,padx=30,pady=30)
    popup.title("Add Dispached Info")
    data = session.query(Sell).filter(Sell.id==idSales).first()

    sale_dispatched_info_var = IntVar()
    sale_dispatched_info_Entry = ttk.Checkbutton(popup,text="Dispatched",variable=sale_dispatched_info_var)
    sale_dispatched_info_Entry.grid(row=0,column=0,pady=10,sticky="w")


    sale_dispatched_date_Label = ttk.Label(popup,text="Dispatched Date",anchor="w")
    sale_dispatched_date_Label.grid(row=1,column=0,pady=10,sticky="w")
    sale_dispatched_date_Entry = DateEntry(popup, selectmode = 'day', date_pattern="dd-mm-yyyy",state="disabled")
    sale_dispatched_date_Entry.grid(row=1,column=1,padx=10,pady=10)

    if(sale_dispatched_info_var):
        sale_dispatched_date_Entry.config(state="normal")

    def addDispatched():
        data = session.query(Sell).filter(Sell.id==idSales).first()
        if(sale_dispatched_info_var):
            data.Dispatched = sale_dispatched_info_var.get()
            data.DispatchDate = sale_dispatched_date_Entry.get_date()
            session.commit()
            status_sale_dispatched["text"] = "Status: Dispatched Added Successfully"
            status_sale_dispatched.config(background='green',foreground='white')
            status_sale_dispatched.after(10000,lambda:status_sale_dispatched.config(text="Status:",background='white',foreground='black'))

    sale_dispatched_btn = ttk.Button(popup,text="Update Dispatched Status",command=addDispatched)
    sale_dispatched_btn.grid(row=2,column=0,pady=20,sticky="w")

    status_sale_dispatched = ttk.Label(popup,text="Status: ",anchor="w")
    status_sale_dispatched.config(background='white',foreground='black')
    status_sale_dispatched.grid(row=3,column=0,sticky="we",pady=5,columnspan=3)

def purchaseListInsert(items):
    purchaseList.delete(0,END)
    i=1
    for item in items:
        purchaseList.insert(i,item)
        i+=1

def getPurchase():
    aFrom = auctionDateEntryFilter.get_date()
    aTo = auctionDateToEntryFilter.get_date()
    f = firmNameVarFilter.get()
    m = markingidPurchaseFilter.get()
    c = coldVarFilter.get()
    all = getAllCheckVar.get()
    if(all==1):
        data = session.query(Purchase).all()
        purchaseListInsert(data)
        
    
    elif(f!=partyNameOptions[0] and m!="" and c!=coldFacilityOptions[0]):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.FirmName==int(f.split(" ")[0]),
                                              Purchase.MarkingID==m,
                                              Purchase.Box>0,
                                              Purchase.coldfacility_id==int(c.split(" ")[0]))
        purchaseListInsert(data)
       
    
    elif(f!=partyNameOptions[0] and c!=coldFacilityOptions[0]):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.FirmName==int(f.split(" ")[0]),
                                              Purchase.Box>0,
                                              Purchase.coldfacility_id==int(c.split(" ")[0]))
        purchaseListInsert(data)
       
    
    elif(m!="" and c!=coldFacilityOptions[0]):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.MarkingID==m,
                                              Purchase.Box>0,
                                              Purchase.coldfacility_id==int(c.split(" ")[0]))
        purchaseListInsert(data)
       
    
    elif(f!=coldFacilityOptions[0] and m!=""):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.MarkingID==m,
                                              Purchase.Box>0,
                                              Purchase.FirmName==int(f.split(" ")[0]))
        purchaseListInsert(data)
        
    
    elif(f!=coldFacilityOptions[0]):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.Box>0,
                                              Purchase.FirmName==int(f.split(" ")[0]))
        purchaseListInsert(data)
       
    
    elif(m!=""):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.Box>0,
                                              Purchase.MarkingID==m)
        purchaseListInsert(data)

    elif(c!=coldFacilityOptions[0]):
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.Box>0,
                                              Purchase.coldfacility_id==int(c.split(" ")[0]))
        purchaseListInsert(data)
        

    else:
        data = session.query(Purchase).filter(Purchase.AuctionDate>=aFrom,Purchase.AuctionDate<=aTo,
                                              Purchase.Box>0)
        purchaseListInsert(data)


    purchase_list_scroll = Scrollbar(frame1,orient='vertical',command=purchaseList.yview)
    purchase_list_scroll.grid(row=7,column=8,sticky="ens")
    purchaseList.configure(yscrollcommand=purchase_list_scroll.set)

def salesListInsert(items):
    salesList.delete(0,END)
    i=1
    # for item in items:
    #     salesList.insert(i,item)
    #     i+=1
    for item in items:
        d = str(item).split(' | ')
        salesList.insert(i,f"{d[0]} | {d[1]} | Sold To: {d[3]} | {d[4]} | {d[5].split(',')[0]}")
        i+=1              

def getSales():
    s = sale_soldToNameVarFilter.get()
    m = sale_markingidPurchaseFilter.get()
    all = sale_getAllCheckVar.get()
    
    if(all==1):
        data = session.query(Sell).all()
        if(data):
            salesListInsert(data)
            return

    elif(s!=defaultPartyOption and m!=""):
        data = session.query(Sell,Purchase).filter(Sell.purchase_id == Purchase.id,
                                                   Sell.SellTo==s,
                                                   Sell.MarkingID==m)
        if(data):
            salesListInsert(data)      
            return
    
    elif(s!=defaultPartyOption):
        data = session.query(Sell,Purchase).filter(Sell.purchase_id == Purchase.id,
                                                   Sell.SellTo==s)
        if(data):
            salesListInsert(data)      
            return
    
    elif(m!=""):
        data = session.query(Sell,Purchase).filter(Sell.purchase_id == Purchase.id,
                                                   Sell.MarkingID==m)
        if(data):
            salesListInsert(data)      
            return
        

    else:
        data = session.query(Sell).filter(Sell.Dispatched == 0)
        if(data):
            salesListInsert(data)
            return

def addColdFacility():
    if(coldFacilityEntry != ""):
        c = ColdFacility(Name=coldFacilityVar.get())                     
        session.add(c)
        session.commit()
        coldFacilityEntry.delete(0,END)
        getColdFacility()
        getColdFacilityOptions()
        # global coldbox,coldboxFilter,coldVar,coldVarFilter
        # coldbox = ttk.OptionMenu(frame1,coldVar,*coldFacilityOptions)
        # coldboxFilter = ttk.OptionMenu(frame1,coldVarFilter,*coldFacilityOptions)
        status3["text"] = "Status: Cold Facility Added successfully"
        status3.config(background='green',foreground='white')
        status3.after(3000,lambda:status3.config(text="Status:",background='white',foreground='black'))
        coldbox['menu'].add_command(label=c, command=_setit(coldVar, c))
        coldboxFilter['menu'].add_command(label=c, command=_setit(coldVarFilter, c))

def deleteColdFacility():
    if(coldFacilityEntry != ""):
        data = session.query(ColdFacility).filter(ColdFacility.id == idCold).first()
        if(data):
            session.delete(data)
            session.commit()
            getColdFacility()
            getColdFacilityOptions()
            global coldbox,coldboxFilter,coldVar,coldVarFilter
            coldbox = ttk.OptionMenu(frame1,coldVar,*coldFacilityOptions)
            coldboxFilter = ttk.OptionMenu(frame1,coldVarFilter,*coldFacilityOptions)
            status3["text"] = "Status: Cold Facility deleted successfully"
            status3.config(background='red',foreground='white')
            status3.after(3000,lambda:status3.config(text="Status:",background='white',foreground='black'))

def editColdFacility():
    if(coldFacilityVar.get() != ""):       
        data = session.query(ColdFacility).filter(ColdFacility.id == idCold).first()
        if(data):
            data.Name = coldFacilityVar.get()
            session.commit()
            clearPurchase()
            getColdFacility()
            status3["text"] = "Status: Cold Facility Edit successfully"
            status3.config(background='green',foreground='white')
            status3.after(10000,lambda:status3.config(text="Status:",background='white',foreground='black'))

def getColdFacility():
    data = session.query(ColdFacility).all()
    coldListInsert(data)

def coldListInsert(items):
    coldFacilityList.delete(0,END)
    i=1
    for item in items:
        coldFacilityList.insert(i,item)
        i+=1
    
def addpartyName():
    if(partyNameEntry != ""):
        p = PartyName(Name=partyNameVar.get())                     
        session.add(p)
        session.commit()
        partyNameEntry.delete(0,END)
        getpartyName()
        getpartyNameOptions()
        # global coldbox,coldboxFilter,coldVar,coldVarFilter
        # coldbox = ttk.OptionMenu(frame1,coldVar,*partyNameOptions)
        # coldboxFilter = ttk.OptionMenu(frame1,coldVarFilter,*partyNameOptions)
        status4["text"] = "Status: Party Name Added successfully"
        status4.config(background='green',foreground='white')
        status4.after(3000,lambda:status4.config(text="Status:",background='white',foreground='black'))
        firmNameEntry['menu'].delete(0, 'end')
        firmNameEntryFilter['menu'].delete(0, 'end')
        sale_soldToNameFilter['menu'].delete(0, 'end')
        data = session.query(PartyName).order_by(asc(PartyName.Name)).all()
        for item in data:
            firmNameEntry['menu'].add_command(label=item, command=_setit(firmNameVar, item))
            firmNameEntryFilter['menu'].add_command(label=item, command=_setit(firmNameVarFilter, item))
            sale_soldToNameFilter['menu'].add_command(label=item, command=_setit(sale_soldToNameVarFilter, item))

def deletepartyName():
    if(partyNameEntry != ""):
        data = session.query(PartyName).filter(PartyName.id == idParty).first()
        if(data):
            session.delete(data)
            session.commit()
            getpartyName()
            getpartyNameOptions()
            global coldbox,coldboxFilter,coldVar,coldVarFilter          # Remains edit
            coldbox = ttk.OptionMenu(frame1,coldVar,*partyNameOptions)
            coldboxFilter = ttk.OptionMenu(frame1,coldVarFilter,*partyNameOptions)
            status4["text"] = "Status: Party deleted successfully"
            status4.config(background='red',foreground='white')
            status4.after(3000,lambda:status4.config(text="Status:",background='white',foreground='black'))

def editpartyName():
    if(partyNameVar.get() != ""):       
        data = session.query(PartyName).filter(PartyName.id == idParty).first()
        if(data):
            data.Name = partyNameVar.get()
            session.commit()
            clearPurchase()                         #  To be understand
            getpartyName()
            status4["text"] = "Status: Party Edit successfully"
            status4.config(background='green',foreground='white')
            status4.after(10000,lambda:status4.config(text="Status:",background='white',foreground='black'))

def getpartyName():
    data = session.query(PartyName).all()
    partyListInsert(data)

def partyListInsert(items):
    partyNameList.delete(0,END)
    i=1
    for item in items:
        partyNameList.insert(i,item)
        i+=1
    
def set_row_background_color(sheet, row_index, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in sheet[row_index]:
        cell.fill = fill

def exportWeightData():
    sheet_name = f"Pending_Weight_{random.randint(1,1000)}"
    try:
        workbook = openpyxl.Workbook()
        workbook.active.title = sheet_name
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        sheet.append(["S.No","Auction Date","Purchased From","Marking ID","BOX","Auction Rate","Net Rate","Net weight","Cold Facility"])        

        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 13
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 10

        all_data = weight_tree.get_children()
        for i, item in enumerate(all_data):
            values = weight_tree.item(item)["values"]
            SNocell = sheet.cell(row=i+2, column=1)
            SNocell.value = i+1
            SNocell.alignment = Alignment(horizontal="left")

            for j, value in enumerate(values):
                cell = sheet.cell(row=i+2, column=j+2)
                cell.value = value
                cell.alignment = Alignment(horizontal="left")

        set_row_background_color(sheet, 1, "FFFF00")
        workbook.save(f"{sheet_name}_data.xlsx")
        status_weight["text"] = f"Status: Data is exported in {sheet_name}_data.xlsx file."
        status_weight.config(background='green',foreground='white')
        status_weight.after(3000,lambda:status_weight.config(text="Status:",background='white',foreground='black'))

    except PermissionError as e:
        status3["text"] = f"Status: {sheet_name}_data.xlsx file is opened somewhere, close the file to export!"
        status3.config(background='red',foreground='white')
        status3.after(3000,lambda:status3.config(text="Status:",background='white',foreground='black'))


def convert_to_excel():
    data = session.query(ColdFacility).filter(ColdFacility.id == idCold).first()
    sheet_name = data.Name
    try:
        workbook = openpyxl.Workbook()
        workbook.active.title = sheet_name
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        sheet.append(["S.No","Auction Date","Purchased From","Marking ID","BOX","Auction Rate","Weight","Sold To","Cold"])        

        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 16
        sheet.column_dimensions['D'].width = 13
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 13
        sheet.column_dimensions['G'].width = 13
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 10

        all_cold = all_coldFacility_data.get_children()
        for i, item in enumerate(all_cold):
            values = all_coldFacility_data.item(item)["values"]
            SNocell = sheet.cell(row=i+2, column=1)
            SNocell.value = i+1
            SNocell.alignment = Alignment(horizontal="left")

            for j, value in enumerate(values):
                cell = sheet.cell(row=i+2, column=j+2)
                cell.value = value
                cell.alignment = Alignment(horizontal="left")

        set_row_background_color(sheet, 1, "FFFF00")
        workbook.save(f"{sheet_name}_data.xlsx")
        status3["text"] = f"Status: Data is exported in {sheet_name}_data.xlsx file."
        status3.config(background='green',foreground='white')
        status3.after(3000,lambda:status3.config(text="Status:",background='white',foreground='black'))

    except PermissionError as e:
        status3["text"] = f"Status: {sheet_name}_data.xlsx file is opened somewhere, close the file to export!"
        status3.config(background='red',foreground='white')
        status3.after(3000,lambda:status3.config(text="Status:",background='white',foreground='black'))

def getreport():
    c=session.query(ColdFacility).all()
    report.delete(0,END)
    for item in c:
        # print("cold id",item.id)
        box_in_cold = session.query(Purchase).filter(Purchase.coldfacility_id == item.id)
        total_box = 0
        for lis in box_in_cold:
            # if(lis.sells):
            #     print("list sells ",lis.sells)
            #     full_string = str(lis.sells)
            #     dispatch_info = full_string.split(" | ")[5].split(": ")[1][:-1]
            #     print("dispatch ",dispatch_info)
            #     if(dispatch_info == 'False'):
            #         sold_box = int(full_string.split(" | ")[4].split(": ")[1])
            #         print("sold box",sold_box)
            #         total_box = lis.Box + sold_box
            
            total_box += lis.Box
    
            # total_box = lis.Box + lis.sells
        data = session.query(Purchase,Sell).filter(Purchase.id==Sell.purchase_id,
                                                   Sell.Dispatched==0,
                                                   Purchase.coldfacility_id==item.id)
        # print("data ",data.all())
        if data:
            for i in data.all():
                l = str(i).split(' | ')
                # print("lllllllll",l)
                # print("ghhh",int(l[12].split(":")[1]))
                if int(l[12].split(":")[1]):
                    total_box += int(l[12].split(":")[1])

        report.insert(END,f"Total No of Box in {item.Name} = {total_box}")

    unsold = session.query(Purchase).filter(Purchase.Box > 0)
    box_unsold = 0
    for item in unsold:
        box_unsold += item.Box
    unsold_box['text'] = f"Total Unsold Box = {box_unsold}"

    sold = session.query(Sell).all()
    box_sold = 0
    for item in sold:
        box_sold += item.Box
    total_sold_box['text'] = f"Total Sold Box = {box_sold}"

    weight_pen = session.query(Purchase).filter(Purchase.weight < 1)
    pending_weight['text'] = f"Total Pending weight LOTS = {weight_pen.count()}"
    
    

# FRAME NO. 1  ================================================================ 
# ============================================================================= 
# ============================================================================= 
frame1 = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")

# All Fields
Label(frame1,text="Add new purchase",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=0,column=0,sticky="we",columnspan=8,pady=10,ipady=3,ipadx=5)

idPurchase = IntVar()
firmNameVar = StringVar()
markingidPurchase = StringVar()
boxVar = IntVar()
auctionRateVar = StringVar()
coldVar = StringVar()
defaultPartyOption = "0 None"
rateVar = DoubleVar()
weightVar = DoubleVar(value=0)
defaultColdFacilityOption = "0 None"
firmNameVarFilter = StringVar()
markingidPurchaseFilter = StringVar()
coldVarFilter = StringVar()
getAllCheckVar = IntVar()
sale_soldToNameVarFilter = StringVar()
partyNameOptions = [defaultPartyOption]
coldFacilityOptions = [defaultColdFacilityOption]

auctionDateLabel = ttk.Label(master=frame1,text="Auction Date")
auctionDateLabel.grid(row=1,column=0)
auctionDateEntry = DateEntry(master=frame1, selectmode = 'day', date_pattern="dd-mm-yyyy")
auctionDateEntry.grid(row=2,column=0)

firmNameLabel = ttk.Label(master=frame1,text="Firm Name")
firmNameLabel.grid(row=1,column=1)


getpartyNameOptions()

# firmNameEntry = ttk.OptionMenu(frame1,firmNameVar,defaultPartyOption,*partyNameOptions)
firmNameEntry =  AutocompleteCombobox(frame1, textvariable=firmNameVar, completevalues=partyNameOptions)
clear_partyName_option()
firmNameEntry.grid(row=2,column=1,padx=3)

markingIdLabel = ttk.Label(master=frame1,text="Marking Id")
markingIdLabel.grid(row=1,column=2)
markingIdEntry = ttk.Entry(master=frame1, textvariable=markingidPurchase)
markingIdEntry.grid(row=2,column=2,padx=3)

boxLabel = ttk.Label(master=frame1,text="Box")
boxLabel.grid(row=1,column=3)
boxEntry = ttk.Entry(master=frame1, textvariable=boxVar)
boxEntry.grid(row=2,column=3,padx=3)

auctionRateLabel = ttk.Label(master=frame1,text="Auction Rate")
auctionRateLabel.grid(row=1,column=4)
auctionRateEntry = ttk.Entry(master=frame1, textvariable=auctionRateVar)
auctionRateEntry.grid(row=2,column=4,padx=3)

def update_rateVar(*args):
    if auctionRateVar.get() !="":
        rateVar.set(value=round(float(auctionRateVar.get())+(float(auctionRateVar.get())*0.025),1))

rateLabel = ttk.Label(master=frame1,text="Rate")
rateLabel.grid(row=1,column=5)
rateEntry = ttk.Label(master=frame1, textvariable=rateVar,background="white")
rateEntry.grid(row=2,column=5,padx=3,sticky="we")
auctionRateVar.trace("w",update_rateVar)

weightLabel = ttk.Label(master=frame1,text="Net Weight")
weightLabel.grid(row=1,column=6)
weightEntry = ttk.Entry(master=frame1, textvariable=weightVar)
weightEntry.grid(row=2,column=6,padx=3)

coldLabel = ttk.Label(master=frame1,text="Cold Facility")
coldLabel.grid(row=1,column=7,padx=3,sticky="WE")

getColdFacilityOptions()
coldbox = ttk.OptionMenu(frame1,coldVar,defaultColdFacilityOption,*coldFacilityOptions)
clear_coldFacility_option()
coldbox.grid(row=2,column=7,padx=3,sticky="WE")

# Buttons
addmultiplePurchaseBtn = ttk.Button(master=frame1, text="Add Multiple Purchases", command=add_multiple_purchases)
addmultiplePurchaseBtn.grid(row=3,column=5,sticky="WE",pady=10,padx=3)
addmultiplePurchaseBtn = ttk.Button(master=frame1, text="Refresh", command=getPurchase)
addmultiplePurchaseBtn.grid(row=3,column=6,sticky="WE",pady=10,padx=3)
addPurchaseBtn = ttk.Button(master=frame1,text="Add Purchase",command=addPurchase)
addPurchaseBtn.grid(row=3,column=1,sticky="WE",pady=10,padx=3)

deletePurchaseBtn = ttk.Button(master=frame1,text="Delete Purchase",command=deletePurchase)
deletePurchaseBtn.grid(row=3,column=2,sticky="WE",padx=3)

editPurchaseBtn = ttk.Button(master=frame1,text="Update Purchase",command=editPurchase)
editPurchaseBtn.grid(row=3,column=3,sticky="WE",padx=3)

clearPurchaseBtn = ttk.Button(master=frame1,text="Clear Fields",command=clearPurchase)
clearPurchaseBtn.grid(row=3,column=4,sticky="WE",padx=3)

#============================================================================================

Label(frame1,text="Search Data and Add Filters",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=4,column=0,sticky="we",columnspan=8,pady=20,ipady=3,ipadx=5)

# filters
auctionDateLabelFilter = ttk.Label(master=frame1,text="Auction Date From")
auctionDateLabelFilter.grid(row=5,column=0,sticky="WE",padx=3)
auctionDateEntryFilter = DateEntry(master=frame1, selectmode = 'day', date_pattern="dd-mm-yyyy")
auctionDateEntryFilter.grid(row=6,column=0,sticky="WE",padx=3)

Label(master=frame1,text="Auction Date To").grid(row=5,column=1,sticky="WE",padx=3)
auctionDateToEntryFilter = DateEntry(master=frame1, selectmode = 'day', date_pattern="dd-mm-yyyy")
auctionDateToEntryFilter.grid(row=6,column=1,sticky="WE",padx=3)

firmNameLabelFilter = ttk.Label(master=frame1,text="Firm Name")
firmNameLabelFilter.grid(row=5,column=2,padx=3)
# firmNameEntryFilter = ttk.OptionMenu(frame1,firmNameVarFilter,defaultPartyOption,*partyNameOptions)
firmNameEntryFilter =  AutocompleteCombobox(frame1, textvariable=firmNameVarFilter, completevalues=partyNameOptions)
clear_partyName_option()
firmNameEntry.grid(row=2,column=1,padx=3)
firmNameEntryFilter.grid(row=6,column=2,padx=3)

markingIdLabelFilter = ttk.Label(master=frame1,text="Marking Id")
markingIdLabelFilter.grid(row=5,column=3,padx=3)
markingIdEntryFilter = ttk.Entry(master=frame1,textvariable=markingidPurchaseFilter)
markingIdEntryFilter.grid(row=6,column=3,padx=3)

coldLabelFilter = ttk.Label(master=frame1,text="Cold Facility")
coldLabelFilter.grid(row=5,column=4,sticky="WE",padx=3)

coldboxFilter = ttk.OptionMenu(frame1,coldVarFilter,defaultColdFacilityOption,*coldFacilityOptions)
clear_coldFacility_option()
coldboxFilter.grid(row=6,column=4,sticky="WE",padx=3)

getAllCheckBox = ttk.Checkbutton(frame1,text="GET All Purchases",variable=getAllCheckVar)
getAllCheckBox.grid(row=6,column=5,sticky="WE",padx=3)

purchaseFilterBtn = ttk.Button(master=frame1,text="Filter",command=getPurchase)
purchaseFilterBtn.grid(row=6,column=6,sticky="WE",padx=3)

#==========================================================================================

purchaseList = Listbox(frame1,height=int(h*0.06),width=int(w*0.5),font=font.Font(size=10))
purchaseList.grid(row=7,column=0,sticky="we",columnspan=8,pady=10)
purchaseList.bind('<<ListboxSelect>>',purchaseListSelect)


getPurchase()

addSalesbtn = ttk.Button(master=frame1,text="Sales To",command=addSales,state="disabled")
addSalesbtn.grid(row=8,column=0,sticky="WE",padx=3)

status1 = ttk.Label(master=frame1,text="Status: ",anchor="w",font=('times new roman',12))
status1.config(background='white',foreground='black')
status1.grid(row=9,column=0,columnspan=8,sticky="we",pady=10)

purchase_info = Listbox(master=frame1,height=3,fg="red",font=('times new roman',14))
purchase_info.grid(row=10,column=0,columnspan=8,sticky="we",ipadx=10,ipady=10)

# report frame

report_frame = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")
report_frame.grid(row=1,column=9,rowspan=10,columnspan=2,sticky="nsew",padx=10)
rootLable = ttk.Label(master=report_frame,text="Welecome to Inventory Management System | Select from above options",font=("times new roman",20),anchor="center")
rootLable.grid(row=0,column=0,sticky="we",columnspan=8)

ttk.Label(master=report_frame,text="Report at Glance",font=('arial',24,"bold")).grid(row=2,column=0,sticky="nsew",padx=10,pady=20)
unsold_box = ttk.Label(master=report_frame,text="Total Unsold Box = ",font=('arial',14))
unsold_box.grid(row=8,column=0,columnspan=9,sticky="nsew",pady=5)
total_sold_box = ttk.Label(master=report_frame,text="Total Sold Box = ",font=('arial',14))
total_sold_box.grid(row=9,column=0,sticky="w",pady=5)
pending_weight = ttk.Label(master=report_frame,text="Total Pending weight LOTS = ",font=('arial',14))
pending_weight.grid(row=10,column=0,sticky="w",pady=5)
report = Listbox(master=report_frame,font=('arial',14),relief="flat",width=int(w*0.30),height=int(h*0.075))
report.grid(row=3,column=0,sticky="we")
getreport()
report_refresh_btn = ttk.Button(master=report_frame,text="Refresh Report",command=getreport)
report_refresh_btn.grid(row=9,column=0,ipadx=20,ipady=3,pady=10)

# Pending Weight Lots Report

weight_frame = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")
# weight_frame.grid(row=1,column=9,rowspan=10,columnspan=2,padx=10)

    
def OnWeightSelect(event):
    try:
        index = weight_tree.selection()
        selected_item = weight_tree.item(index)
        item_values = selected_item["values"]

        global updateIndex
        updateIndex=item_values[0] 
        update_weight_btn.config(state="normal")     

    except:
        pass

def open_update_modal():
    # Create a new dialog window
        
    update_dialog = Toplevel(weight_frame)
    update_dialog.title("Update Slot")

    if not updateIndex:
        messagebox.showerror("error", "Please select a slot!", parent=update_dialog)
        update_dialog.destroy()

    ttk.Label(update_dialog, text="Slot ID").grid(row=0, column=0)
    ttk.Label(update_dialog, text=updateIndex).grid(row=0, column=3)
    ttk.Label(update_dialog, text="Weight:").grid(row=1, column=0)
    update_net_weight_entry = ttk.Entry(update_dialog)
    update_net_weight_entry.grid(row=2, column=0, columnspan=5, pady=10, padx=10)

    getColdFacilityOptions()
    ttk.Label(update_dialog, text="Cold Fasicility:").grid(row=1, column=6)

    update_cold_facility_var = StringVar()
    update_cold_facility_dropdown = ttk.OptionMenu(update_dialog,update_cold_facility_var,defaultColdFacilityOption,*coldFacilityOptions)
    # cold_facility_dropdown = ttk.Combobox(dialog, textvariable=cold_facility_var)
    # cold_facility_dropdown['values'] = ('Cold A', 'Cold B', 'Cold C')  # Replace with your cold facilities
    update_cold_facility_dropdown.grid(row=2, column=5, columnspan=5, pady=10, padx=10)

    def update_weight():

        try:
            data = session.query(Purchase).filter(Purchase.id == updateIndex).first()
            if(data):
                data.weight=update_net_weight_entry.get()
                if len(update_cold_facility_var.get().split(" | "))>1:
                    data.coldfacility_id=int(update_cold_facility_var.get().split(" | ")[0])
                session.commit()
                getPendingWeight()
                messagebox.showinfo("info", "Data updated!", parent=update_dialog)
            else:
                messagebox.showerror("error", "Data not found!", parent=update_dialog)
        except Exception as e:
            messagebox.showerror("error", "Error occured, Try again!", parent=update_dialog)

        finally:
            update_weight_btn.config(state="disabled")   
            update_dialog.destroy()
        
        
        
    ttk.Button(update_dialog, text="Update", command=update_weight).grid(row=3, column=6, pady=10, padx=10)


ttk.Label(master=weight_frame,text="Pending Weight Lots",font=('arial',24,"bold")).grid(row=0,column=0,sticky="nsew",padx=10,pady=20)

weight_tree = ttk.Treeview(weight_frame, selectmode ='browse',height=int(h*0.1)+3)
weight_tree.grid(row=2,column=0,rowspan=20,columnspan=5,sticky='w',ipadx=20)
weight_tree.bind("<<TreeviewSelect>>", OnWeightSelect)
weight_tree.tag_configure("custom",font=("Aerial",12))

weight_tree["columns"] = ('1','2','3','4','5','6','7','8', '9')

weight_tree.column("#0", width= 60, anchor ='w')
weight_tree.column("1", width= 60, anchor ='w')
weight_tree.column("2", width = 140, anchor ='w')
weight_tree.column("3", width = 200, anchor ='w')
weight_tree.column("4", width = 120, anchor ='w')
weight_tree.column("5", width = 80, anchor ='w')
weight_tree.column("6", width = 120, anchor ='w')
weight_tree.column("7", width = 100, anchor ='w')
weight_tree.column("8", width = 90, anchor ='w')
weight_tree.column("9", width = 80, anchor ='center')

weight_tree.heading("#0", text ="S.No")
weight_tree.heading("1", text ="ID")
weight_tree.heading("2", text ="Auction Date")
weight_tree.heading("3", text ="Purchased From")
weight_tree.heading("4", text ="Marking ID")
weight_tree.heading("5", text ="BOX")
weight_tree.heading("6", text ="Auction Rate")
weight_tree.heading("7", text ="Net Rate")
weight_tree.heading("8", text ="Weight")
weight_tree.heading("9", text ="Cold")
getPendingWeight()
# updateIndex = None

# weight_report.grid(row=3,column=0,sticky="we")
update_weight_btn = ttk.Button(master=weight_frame,text="Refresh",command=getPendingWeight)
update_weight_btn.grid(row=22,column=3,ipadx=20,ipady=3,pady=10)
update_weight_btn = ttk.Button(master=weight_frame,text="Update",command=open_update_modal, state="disabled")
update_weight_btn.grid(row=22,column=4,ipadx=20,ipady=3,pady=10)
export_weight_btn = ttk.Button(master=weight_frame,text="Export",command=exportWeightData)
export_weight_btn.grid(row=22,column=5,ipadx=20,ipady=3,pady=10)
status_weight = ttk.Label(master=weight_frame,text="Status: ",anchor="w",font=('times new roman',12))
status_weight.config(background='white',foreground='black')
status_weight.grid(row=22,column=0,columnspan=3,sticky="we",pady=10)



# FRAME NO. 2 ==============================================================================
#===========================================================================================
#===========================================================================================

frame2 = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")

idSales = IntVar()

Label(frame2,text="All Sales | OR add Filters",font=("times new roman",16),anchor="w",bg="#696562",fg="white").grid(row=0,column=0,sticky="we",columnspan=10,pady=10,ipady=3,ipadx=5)


sale_soldToNameLabelFilter = ttk.Label(master=frame2,text="Sold To Name",anchor="center")
sale_soldToNameLabelFilter.grid(row=1,column=0,sticky="NSEW")
# sale_soldToNameFilter = ttk.OptionMenu(frame2,sale_soldToNameVarFilter,defaultPartyOption,*partyNameOptions)
sale_soldToNameFilter =  AutocompleteCombobox(frame2, textvariable=sale_soldToNameVarFilter, completevalues=partyNameOptions)
clear_partyName_option()
sale_soldToNameFilter.grid(row=2,column=0,sticky="WE",pady=10)

sale_markingidPurchaseFilter = StringVar()
sale_markingIdLabelFilter = ttk.Label(master=frame2,text="Marking Id")
sale_markingIdLabelFilter.grid(row=1,column=1,sticky="WE",padx=30)
sale_markingIdEntryFilter = ttk.Entry(master=frame2,textvariable=sale_markingidPurchaseFilter)
sale_markingIdEntryFilter.grid(row=2,column=1,sticky="WE",padx=30,pady=10)

# sale_coldVarFilter = StringVar()
# sale_coldLabelFilter = Label(master=frame2,text="Cold Facilit")
# sale_coldLabelFilter.grid(row=1,column=2)

# sale_coldboxFilter = OptionMenu(frame2,coldVarFilter,*coldFacilityOptions)
# sale_coldVarFilter.set(coldFacilityOptions[0])
# sale_coldboxFilter.grid(row=2,column=2,sticky="WE")

sale_getAllCheckVar = IntVar()
sale_getAllCheckBox = ttk.Checkbutton(master=frame2,text="GET All Sales",variable=sale_getAllCheckVar)
sale_getAllCheckBox.grid(row=2,column=2)
 

# Buttons
sale_FilterBtn = ttk.Button(master=frame2,text="Filter",command=getSales)
sale_FilterBtn.grid(row=2,column=3,sticky="WE")

salesList = Listbox(frame2,height=int(h*0.1),width=int(w*0.5),font=font.Font(size=10))
salesList.grid(row=3,column=0,sticky="we",columnspan=10)
salesList.bind('<<ListboxSelect>>',salesListSelect)

getSales()


def dispatchedSalesMultiple():
    popup=Toplevel(master=frame2,padx=30,pady=30)
    popup.title("Add Dispached Info")
    # data = session.query(Sell).filter(Sell.id==idSales).first()

    sale_dispatched_info_var = IntVar()
    sale_dispatched_info_Entry = ttk.Checkbutton(popup,text="Dispatched",variable=sale_dispatched_info_var)
    sale_dispatched_info_Entry.grid(row=0,column=0,pady=10,sticky="w")


    sale_dispatched_date_Label = ttk.Label(popup,text="Dispatched Date",anchor="w")
    sale_dispatched_date_Label.grid(row=1,column=0,pady=10,sticky="w")
    sale_dispatched_date_Entry = DateEntry(popup, selectmode = 'day', date_pattern="dd-mm-yyyy",state="disabled")
    sale_dispatched_date_Entry.grid(row=1,column=1,padx=10,pady=10)

    if(sale_dispatched_info_var):
        sale_dispatched_date_Entry.config(state="normal")

    def addDispatched():
        salesItems = salesList.curselection()
        for index in salesItems:
            text = str(salesList.get(index))
            datalist = text.split(sep=" | ")
            idSalesSelect=int(datalist[0].split(': ')[1])
            data = session.query(Sell).filter(Sell.id==idSalesSelect).first()
            if(sale_dispatched_info_var):
                data.Dispatched = sale_dispatched_info_var.get()
                data.DispatchDate = sale_dispatched_date_Entry.get_date()
        session.commit()
        status_sale_dispatched["text"] = "Status: Dispatched Added Successfully"
        status_sale_dispatched.config(background='green',foreground='white')
        status_sale_dispatched.after(10000,lambda:status_sale_dispatched.config(text="Status:",background='white',foreground='black'))

    sale_dispatched_btn = ttk.Button(popup,text="Update Dispatched Status",command=addDispatched)
    sale_dispatched_btn.grid(row=2,column=0,pady=20,sticky="w")

    status_sale_dispatched = ttk.Label(popup,text="Status: ",anchor="w")
    status_sale_dispatched.config(background='white',foreground='black')
    status_sale_dispatched.grid(row=3,column=0,sticky="we",pady=5,columnspan=3)


def deleteSalesMultiple():
    try:
        salesItems = salesList.curselection()
        for index in salesItems:
            text = str(salesList.get(index))
            datalist = text.split(sep=" | ")
            idSalesSelect=int(datalist[0].split(': ')[1])
            data = session.query(Sell).filter(Sell.id == idSalesSelect).first()
            if(data):
                session.delete(data)
            else:
                messagebox.showerror("error", "Data not found", parent=frame2)
        session.commit()
        getSales()
        sale_status["text"] = "Status: Sales deleted successfully"
        sale_status.config(background='green',foreground='white')
        sale_status.after(10000,lambda:sale_status.config(text="Status:",background='white',foreground='black'))
    except:
        messagebox.showerror("error", "Error, Try again!", parent=frame2)


sale_deleteBtn=ttk.Button(master=frame2,text="Delete Sale",command=deleteSales)
sale_deleteBtn.grid(row=4,column=0,pady=15,sticky="WE",padx=5)

# sale_editBtn=ttk.Button(master=frame2,text="Edit Sale",command=editSales)
# sale_editBtn.grid(row=4,column=1)

sale_dispatchedBtn=ttk.Button(master=frame2,text="Enter Dispached INFO",command=dispatchedSales)
sale_dispatchedBtn.grid(row=4,column=1,sticky="WE",padx=30)

sale_refreshbtn=ttk.Button(master=frame2,text="Refresh",command=getSales)
sale_refreshbtn.grid(row=4,column=2,pady=15,sticky="WE",padx=5)

def changeSalesMultiple():
    salesList.config(selectmode=MULTIPLE)
    sale_deleteBtn.config(command=deleteSalesMultiple)
    sale_dispatchedBtn.config(command=dispatchedSalesMultiple)

def changeSalesSingle():
    salesList.config(selectmode=SINGLE)
    sale_deleteBtn.config(command=deleteSales)
    sale_dispatchedBtn.config(command=dispatchedSales)

sale_multiplebtn=ttk.Button(master=frame2,text="Select Multiple",command=changeSalesMultiple)
sale_multiplebtn.grid(row=4,column=3,pady=15,sticky="WE",padx=5)

sale_singlebtn=ttk.Button(master=frame2,text="Select Single",command=changeSalesSingle)
sale_singlebtn.grid(row=4,column=4,pady=15,sticky="WE",padx=5)

sale_status = ttk.Label(master=frame2,text="Status: ",anchor="w",font=('times new roman',12))
sale_status.config(background='white',foreground='black')
sale_status.grid(row=5,column=0,columnspan=10,sticky="we",pady=10)

sale_info = Listbox(frame2,height=3,fg="red",font=('times new roman',14),relief="groove")
sale_info.grid(row=6,column=0,columnspan=10,sticky="we",ipady=10)


# FRAME NO. 3==============================================================================
# =========================================================================================
# =========================================================================================

frame3 = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")

# Variables
Label(master=frame3,text="Add New Cold Facility",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=0,column=0,sticky="we",columnspan=3,pady=10,ipady=3,ipadx=5,padx=5)
Label(master=frame3,text="Total LOTS in Cold Facility",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=0,column=3,sticky="we",columnspan=3,pady=10,ipady=3,ipadx=5,padx=5)
# ttk.Label(master=frame3,text="Add New Cold Facility").grid(row=0,column=0)

idCold = IntVar()

coldFacilityVar = StringVar()
coldFacilityLable = ttk.Label(master=frame3,text="Cold Facility Name: ")
coldFacilityLable.grid(row=1,column=0)
coldFacilityEntry = ttk.Entry(master=frame3,textvariable=coldFacilityVar)
coldFacilityEntry.grid(row=1,column=1)

# Buttons
addColdFacilityBtn = ttk.Button(master=frame3,text="Add Cold facility",command=addColdFacility)
addColdFacilityBtn.grid(row=1,column=2,sticky="WE",padx=15)

deleteColdFacilityBtn = ttk.Button(master=frame3,text="Delete Cold facility",command=deleteColdFacility)
deleteColdFacilityBtn.grid(row=2,column=2,sticky="WE",padx=15,ipadx=5)

editColdFacilityBtn = ttk.Button(master=frame3,text="Edit Cold facility",command=editColdFacility)
editColdFacilityBtn.grid(row=3,column=2,sticky="WE",padx=15)

coldFacilityList = Listbox(master=frame3,height=int(h*0.1)+3)
coldFacilityList.grid(row=2,column=0,sticky="WE",columnspan=2,rowspan=5)
coldFacilityList.bind('<<ListboxSelect>>',coldListSelect)

getColdFacility()

status3 = ttk.Label(master=frame3,text="Status: ",anchor="w",font=('times new roman',12))
status3.config(background='white',foreground='black')
status3.grid(row=28,column=0,columnspan=5,sticky="we",pady=10)

# label_text_head = "{:^15} | {:^25} | {:^15} | {:^10} | {:^15} | {:^10} | {:^25} | {:^10}"
# label_text_head = label_text_head.format('Auction Date','Purchased From','Marking ID','BOX','Auction Rate','Weight','Sold To','Cold')
# ttk.Label(master=frame3,text=label_text_head,font=('times new roman',12)).grid(row=1,column=3,sticky="w")


# all_coldFacility_data = Listbox(master=frame3,height=15,width=100,font=('times new roman',12),relief="groove")
# all_coldFacility_data.grid(row=1,column=3,rowspan=14,sticky="we",ipadx=10,ipady=10)

all_coldFacility_data = ttk.Treeview(frame3, selectmode ='browse',height=int(h*0.1)+3)
all_coldFacility_data.grid(row=1,column=3,rowspan=25,sticky='w',ipadx=20)

all_coldFacility_data.tag_configure("custom",font=("Aerial",12))

all_coldFacility_data["columns"] = ('1','2','3','4','5','6','7','8')

all_coldFacility_data.column("#0", width= 60, anchor ='w')
all_coldFacility_data.column("1", width = 110, anchor ='w')
all_coldFacility_data.column("2", width = 120, anchor ='w')
all_coldFacility_data.column("3", width = 100, anchor ='w')
all_coldFacility_data.column("4", width = 70, anchor ='w')
all_coldFacility_data.column("5", width = 100, anchor ='w')
all_coldFacility_data.column("6", width = 90, anchor ='w')
all_coldFacility_data.column("7", width = 110, anchor ='w')
all_coldFacility_data.column("8", width = 70, anchor ='center')

all_coldFacility_data.heading("#0", text ="S.No")
all_coldFacility_data.heading("1", text ="Auction Date")
all_coldFacility_data.heading("2", text ="Purchased From")
all_coldFacility_data.heading("3", text ="Marking ID")
all_coldFacility_data.heading("4", text ="BOX")
all_coldFacility_data.heading("5", text ="Auction Rate")
all_coldFacility_data.heading("6", text ="Weight")
all_coldFacility_data.heading("7", text ="Sold To")
all_coldFacility_data.heading("8", text ="Cold")

exportFile = ttk.Button(master=frame3,text="Export",command=convert_to_excel)
exportFile.grid(row=27,column=3,sticky="E",pady=10)


# FRAME NO. 4==============================================================================
# =========================================================================================
# =========================================================================================

frame4 = Frame(master=root,padx=10,pady=10,borderwidth=3,relief="groove")

# Variables
Label(master=frame4,text="Add New Party Name",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=0,column=0,sticky="we",columnspan=3,pady=10,ipady=3,ipadx=5,padx=5)
Label(master=frame4,text="Total Purachase & Sales of Party",font=("times new roman",13),anchor="w",bg="#696562",fg="white").grid(row=0,column=3,sticky="we",columnspan=3,pady=10,ipady=3,ipadx=5,padx=5)
# ttk.Label(master=frame4,text="Add New Cold Facility").grid(row=0,column=0)

idParty = IntVar()

partyNameVar = StringVar()
partyNameLable = ttk.Label(master=frame4,text="Party Name: ")
partyNameLable.grid(row=1,column=0)
partyNameEntry = ttk.Entry(master=frame4,textvariable=partyNameVar)
partyNameEntry.grid(row=1,column=1)

# Buttons
addpartyNameBtn = ttk.Button(master=frame4,text="Add Party",command=addpartyName)
addpartyNameBtn.grid(row=1,column=2,sticky="WE",padx=15)

deletepartyNameBtn = ttk.Button(master=frame4,text="Delete Party",command=deletepartyName)
deletepartyNameBtn.grid(row=2,column=2,sticky="WE",padx=15,ipadx=5)

editpartyNameBtn = ttk.Button(master=frame4,text="Edit Party",command=editpartyName)
editpartyNameBtn.grid(row=3,column=2,sticky="WE",padx=15)

partyNameList = Listbox(master=frame4,height=int(h*0.1)+5)
partyNameList.grid(row=2,column=0,sticky="WE",columnspan=2,rowspan=5)
partyNameList.bind('<<ListboxSelect>>',partyListSelect)

getpartyName()

status4 = ttk.Label(master=frame4,text="Status: ",anchor="w",font=('times new roman',12))
status4.config(background='white',foreground='black')
status4.grid(row=17,column=0,columnspan=5,sticky="we",pady=10)

all_partyName_data = Listbox(master=frame4,height=int(h*0.1)+5,width=int(w*0.3)+10,font=('times new roman',12),relief="groove")
all_partyName_data.grid(row=1,column=3,rowspan=15,sticky="we",ipadx=10,ipady=10)

#==================================================================================================
root.mainloop()